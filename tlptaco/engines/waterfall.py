"""
Waterfall engine: computes waterfall metrics from eligibility table.
"""
from tlptaco.config.schema import WaterfallConfig, EligibilityConfig
from tlptaco.db.runner import DBRunner
from tlptaco.utils.logging import get_logger
from tlptaco.sql.generator import SQLGenerator
from tlptaco.io.writer import write_dataframe
import os

class WaterfallEngine:
    def __init__(self, cfg: WaterfallConfig, runner: DBRunner, logger=None):
        self.cfg = cfg
        self.runner = runner
        self.logger = logger or get_logger("waterfall")

    def num_steps(self) -> int:
        """Return the number of grouping tasks that run() will execute."""
        return len(self.cfg.count_columns)

    def _extract_check_columns(self, conditions_cfg) -> list:
        """
        Flatten and sort all check names from conditions config.
        """
        cols = []
        # main BA
        for chk in conditions_cfg.main.BA:
            cols.append(chk.name)
        # main others
        if conditions_cfg.main.others:
            for templ, checks in conditions_cfg.main.others.items():
                for chk in checks:
                    cols.append(chk.name)
        # channel templates
        for tmpl_cfg in conditions_cfg.channels.values():
            for chk in tmpl_cfg.BA:
                cols.append(chk.name)
            if tmpl_cfg.others:
                for templ, checks in tmpl_cfg.others.items():
                    for chk in checks:
                        cols.append(chk.name)
        # sort by trailing numeric
        return sorted(cols, key=lambda x: int(x.rsplit('_', 1)[-1]))

    def run(self, eligibility_engine, progress=None):
        """
        Compute waterfall metrics entirely in-database, then export report.
        """
        # Compute waterfall for each group of identifier columns
        elig_cfg = eligibility_engine.cfg  # type: EligibilityConfig
        # Build grouping definitions
        groups = []
        for item in self.cfg.count_columns:
            if isinstance(item, str):
                cols = [item]
                grp_name = item.split('.')[-1]
            else:
                cols = item
                grp_name = '_'.join([c.split('.')[-1] for c in cols])
            groups.append({'name': grp_name, 'cols': cols})
        # Precompute check column order
        check_cols = self._extract_check_columns(elig_cfg.conditions)
        # Template generator
        templates_dir = os.path.join(os.path.dirname(__file__), '..', 'sql', 'templates')
        gen = SQLGenerator(templates_dir)
        # Ensure output directory exists
        out_dir = self.cfg.output_directory
        os.makedirs(out_dir, exist_ok=True)
        import pandas as pd
        # Build list of base checks
        base_checks = [chk.name for chk in elig_cfg.conditions.main.BA]
        if elig_cfg.conditions.main.others:
            for lst in elig_cfg.conditions.main.others.values():
                for chk in lst:
                    base_checks.append(chk.name)
        # Build mapping of check names to SQL logic for labels
        check_map = {}
        # main BA checks
        for chk in elig_cfg.conditions.main.BA:
            check_map[chk.name] = chk.sql
        # main others
        if elig_cfg.conditions.main.others:
            for lst in elig_cfg.conditions.main.others.values():
                for chk in lst:
                    check_map[chk.name] = chk.sql
        # channel checks
        for tmpl_cfg in elig_cfg.conditions.channels.values():
            for chk in tmpl_cfg.BA:
                check_map[chk.name] = chk.sql
            if tmpl_cfg.others:
                for lst in tmpl_cfg.others.values():
                    for chk in lst:
                        check_map[chk.name] = chk.sql
        # Loop through each grouping of identifiers
        for grp in groups:
            name = grp['name']
            uniq_ids = grp['cols']
            # Collect DataFrames for each section
            df_sections = []
            # 1) Base waterfall
            ctx_base = {
                'eligibility_table': elig_cfg.eligibility_table,
                'unique_identifiers': uniq_ids,
                'unique_without_aliases': [u.split('.')[-1] for u in uniq_ids],
                'check_columns': base_checks,
                'pre_filter': None
            }
            sql_base = gen.render('waterfall_full.sql.j2', ctx_base)
            self.logger.info(f'Executing base waterfall SQL for grouping {name}')
            df_base = self.runner.to_df(sql_base)
            if 'stat_name' in df_base.columns:
                df_base = (
                    df_base.set_index('stat_name')
                           .T
                           .reset_index()
                           .rename(columns={'index': 'check_name'})
                )
            df_base['section'] = 'base'
            df_sections.append(df_base)
            # 2) Per-channel waterfalls
            for channel, tmpl_cfg in elig_cfg.conditions.channels.items():
                # gather channel checks
                chks = [chk.name for chk in tmpl_cfg.BA]
                if tmpl_cfg.others:
                    for lst in tmpl_cfg.others.values():
                        for chk in lst:
                            chks.append(chk.name)
                # pre_filter: must pass all base checks
                pf = ' AND '.join([f"c.{c}=1" for c in base_checks])
                # prepare context for channel section
                ctx_ch = {
                    'eligibility_table': elig_cfg.eligibility_table,
                    'unique_identifiers': uniq_ids,
                    'unique_without_aliases': [u.split('.')[-1] for u in uniq_ids],
                    'check_columns': chks,
                    'pre_filter': pf
                }
                sql_ch = gen.render('waterfall_full.sql.j2', ctx_ch)
                self.logger.info(f'Executing waterfall SQL for channel {channel}, grouping {name}')
                df_ch = self.runner.to_df(sql_ch)
                if 'stat_name' in df_ch.columns:
                    df_ch = (
                        df_ch.set_index('stat_name')
                             .T
                             .reset_index()
                             .rename(columns={'index': 'check_name'})
                    )
                df_ch['section'] = channel
                df_sections.append(df_ch)
            # Concatenate all sections into a single DataFrame
            result_df = pd.concat(df_sections, ignore_index=True)
            # Replace check_name labels with the underlying SQL logic
            if 'check_name' in result_df.columns:
                result_df['check_name'] = result_df['check_name'].map(lambda nm: check_map.get(nm, nm))
            # Override column names to full stat names, preserving order
            full_parts = ['unique_drops', 'incremental_drops', 'remaining', 'cumulative_drops']
            cols_new = ['check_name'] + full_parts + ['section']
            if len(result_df.columns) == len(cols_new):
                result_df.columns = cols_new
            else:
                self.logger.warning(
                    f"Expected {len(cols_new)} columns but got {len(result_df.columns)}; skipping rename"
                )
            filename = f"waterfall_{elig_cfg.eligibility_table}_{name}.xlsx"
            out_path = os.path.join(out_dir, filename)
            # Write styled Excel waterfall report
            try:
                import pandas as pd
                from openpyxl.styles import PatternFill, Font
                with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Waterfall')
                    ws = writer.sheets['Waterfall']
                    # Ensure full header names
                    headers = list(result_df.columns)
                    for col_idx, name in enumerate(headers, start=1):
                        ws.cell(row=1, column=col_idx).value = name
                    # Style header row
                    hdr_fill = PatternFill(fill_type='solid', fgColor='FFD700')
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = hdr_fill
                    # Freeze header
                    ws.freeze_panes = 'A2'
                    # Color sections differently
                    sec_col = result_df.columns.get_loc('section') + 1
                    for idx, sec in enumerate(result_df['section'], start=2):
                        if sec == 'base':
                            fill = PatternFill(fill_type='solid', fgColor='DDDDDD')
                        elif sec == 'email':
                            fill = PatternFill(fill_type='solid', fgColor='CCFFFF')
                        elif sec == 'sms':
                            fill = PatternFill(fill_type='solid', fgColor='CCFFCC')
                        else:
                            fill = None
                        if fill:
                            for col in range(1, len(result_df.columns)+1):
                                ws.cell(row=idx, column=col).fill = fill
                # After styling, insert table definitions below the data
                try:
                    rows, cols = result_df.shape
                    # Determine start row for table metadata (header + data + one blank)
                    start_row = rows + 3
                    # Header for tables section
                    ws.cell(row=start_row, column=1).value = 'Source Tables:'
                    from openpyxl.styles import Font
                    ws.cell(row=start_row, column=1).font = Font(bold=True)
                    # Write each table's FROM/JOIN clause
                    for i, tbl in enumerate(elig_cfg.tables, start=1):
                        if i == 1:
                            text = f"FROM {tbl.name} {tbl.alias}"
                            if tbl.where_conditions:
                                text += f" WHERE {tbl.where_conditions}"
                        else:
                            jt = tbl.join_type or 'JOIN'
                            text = f"{jt} {tbl.name} {tbl.alias}"
                            if tbl.join_conditions:
                                text += f" ON {tbl.join_conditions}"
                            if tbl.where_conditions:
                                text += f" WHERE {tbl.where_conditions}"
                        ws.cell(row=start_row + i, column=1).value = text
                except Exception:
                    pass
                # Log file write with row/column counts
                try:
                    rows, cols = result_df.shape
                    self.logger.info(f'Waterfall report saved to {out_path} ({rows} rows, {cols} columns)')
                except Exception:
                    self.logger.info(f'Waterfall report saved to {out_path}')
            except Exception as e:
                self.logger.warning(f'Excel styling failed ({e}), writing plain file')
                write_dataframe(result_df, out_path, fmt='excel')
                # Log plain write with counts
                try:
                    rows, cols = result_df.shape
                    self.logger.info(f'Waterfall report (plain) saved to {out_path} ({rows} rows, {cols} columns)')
                except Exception:
                    self.logger.info(f'Waterfall report (plain) saved to {out_path}')
            # Update progress bar for this grouping
            if progress:
                progress.update("Waterfall")