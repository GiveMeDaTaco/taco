"""
Excel writer for waterfall reports (current + history).

The workbook layout is:

1. Sheet "waterfall" – consolidated report identical to the original format
   (all groups side-by-side).
2. One additional sheet *per group* containing:
      • a full table for the *current* run
      • a full table for the *previous* run within the configured look-back
        window (if no prior data → note displayed instead).

This module has **no legacy compatibility shims** – it is only used by
WaterfallEngine v2 and tests inside this repository.
"""

from __future__ import annotations

from typing import List, Tuple, Dict, Iterable, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ────────────────────────────────────────────────────────────────────────────────
# Constants / helper data -------------------------------------------------------
# ────────────────────────────────────────────────────────────────────────────────


_BASE_TITLES = ["Section", "Template", "#", "Criteria", "Description"]

# Column order & friendly display names for metrics
_METRIC_ORDER: List[Tuple[str, str]] = [
    ("unique_drops", "Drop If Only This Scrub"),
    ("regain", "Regain If No Scrub"),
    ("incremental_drops", "Drop Incremental"),
    ("cumulative_drops", "Drop Cumulative"),
    ("remaining", "Remaining"),
]

_HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# Additional palette colours
_BLOCK_HEADER_FILL = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")  # Cornflower Blue
_HIST_ONLY_FILL   = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light Yellow
_CUR_ONLY_FILL    = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light Mint


# ────────────────────────────────────────────────────────────────────────────────
# Public API --------------------------------------------------------------------
# ────────────────────────────────────────────────────────────────────────────────


def write_waterfall_excel(
    conditions: pd.DataFrame,
    compiled_current: List[Tuple[str, List[Tuple[str, pd.DataFrame]]]],
    output_path: str,
    *,
    previous: Dict[str, Dict[str, Any]] | None = None,
    offer_code: str = "",
    campaign_planner: str = "",
    lead: str = "",
    current_date: str = "",
    starting_pops: Dict[str, int] | None = None,
) -> None:
    """Write the Excel workbook for a Waterfall run.

    Parameters
    ----------
    conditions
        DataFrame indexed by *check_name* with extra descriptive columns.
    compiled_current
        Output of WaterfallEngine for the current run: list of
        ``(group_name, compiled)`` where *compiled* == list of
        ``(section_name, pivoted_df)``.
    output_path
        Destination ``.xlsx`` file.
    previous
        Optional mapping: ``{group_name: compiled_previous}`` following the
        same internal *compiled* structure. Used to render comparison tables
        on individual group sheets.
    starting_pops
        Mapping ``group_name -> starting_population``. If omitted the
        *Starting Population* row is left blank.
    """

    if previous is None:
        previous = {}
    if starting_pops is None:
        starting_pops = {}

    wb = Workbook()
    ws_cons = wb.active
    ws_cons.title = "waterfall"

    _write_header(ws_cons, offer_code, campaign_planner, lead, current_date)
    _write_consolidated_table(ws_cons, conditions, compiled_current, starting_pops)

    # ------------------------------------------------------------------
    # Per-group sheets ---------------------------------------------------
    # ------------------------------------------------------------------
    for group_name, compiled_cur in compiled_current:
        sheet_name = group_name[:31]
        ws_grp = wb.create_sheet(title=sheet_name)

        _write_header(ws_grp, offer_code, campaign_planner, lead, current_date)

        prev_info = previous.get(group_name) if previous else None

        if prev_info:
            hist_date = prev_info.get('date', '')
            compiled_prev = prev_info['compiled']
            prev_start_pop = prev_info.get('start_pop')

            # Comparison table side-by-side
            _write_group_comparison_table(
                ws_grp,
                start_row=2,
                group_name=group_name,
                compiled_current=compiled_cur,
                compiled_historic=compiled_prev,
                conditions=conditions,
                start_pop_current=starting_pops.get(group_name),
                start_pop_historic=prev_start_pop,
                historic_date=hist_date,
            )
        else:
            # Fall back to legacy single table view
            _write_group_table(
                ws_grp,
                start_row=2,
                group_name=group_name,
                compiled=compiled_cur,
                conditions=conditions,
                start_pop=starting_pops.get(group_name),
            )

    wb.save(output_path)

    # Adjust permissions for group-sharing
    from tlptaco.utils.fs import grant_group_rwx
    grant_group_rwx(output_path)


# ────────────────────────────────────────────────────────────────────────────────
# Private helpers ----------------------------------------------------------------
# ────────────────────────────────────────────────────────────────────────────────


def _write_header(ws, offer_code: str, campaign_planner: str, lead: str, current_date: str) -> None:
    """Write the big header row (row 1)."""
    header_txt = f"[[{offer_code}] [CP: {campaign_planner}] [LEAD: {lead}] [DATE: {current_date}]"
    ws["A1"] = header_txt
    ws["A1"].font = Font(size=18)


def _write_consolidated_table(ws, conditions: pd.DataFrame, compiled_groups: List[Tuple[str, List[Tuple[str, pd.DataFrame]]]], starting_pops: Dict[str, int]) -> None:
    """Render the classic side-by-side consolidated sheet (rows begin at 2)."""

    # Column titles – row 2
    for cidx, title in enumerate(_BASE_TITLES, start=1):
        cell = ws.cell(row=2, column=cidx, value=title)
        cell.font = Font(size=12)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    start_col = len(_BASE_TITLES) + 1  # first metric column

    # Metric header rows (rows 2-3)
    col_ptr = start_col
    for group_name, _ in compiled_groups:
        ws.merge_cells(start_row=2, start_column=col_ptr, end_row=2, end_column=col_ptr + len(_METRIC_ORDER) - 1)
        gcell = ws.cell(row=2, column=col_ptr, value=group_name)
        gcell.font = Font(size=10, bold=True)
        gcell.alignment = Alignment(horizontal="center", vertical="center")

        for m_idx, (_, disp) in enumerate(_METRIC_ORDER):
            cell = ws.cell(row=3, column=col_ptr + m_idx, value=disp)
            cell.font = Font(size=9)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = 14

        col_ptr += len(_METRIC_ORDER) + 1  # +1 spacer

    # Build lookup: group -> check_name -> metrics dict
    group_lookup: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for grp, compiled in compiled_groups:
        sub: Dict[str, Dict[str, Any]] = {}
        for _, df in compiled:
            for _, row in df.iterrows():
                sub[row["check_name"]] = row.to_dict()
        group_lookup[grp] = sub

    # Starting population row (row 4)
    row_idx = 4
    ws.cell(row=row_idx, column=5, value="Starting Population").font = Font(size=10, bold=True)

    col_ptr = len(_BASE_TITLES) + _metric_remaining_offset()  # Remaining column within block
    for grp, _ in compiled_groups:
        val = starting_pops.get(grp, "")
        ws.cell(row=row_idx, column=col_ptr, value=val).font = Font(size=10)
        col_ptr += len(_METRIC_ORDER) + 1

    # Conditions rows – begin at row 5
    for _, cond_row in conditions.reset_index().iterrows():
        row_idx += 1
        # Base descriptive columns
        ws.cell(row=row_idx, column=1, value=cond_row["Section"]).font = Font(size=10)
        ws.cell(row=row_idx, column=2, value=cond_row["Template"]).font = Font(size=10)
        ws.cell(row=row_idx, column=3, value=cond_row["#"]).font = Font(size=10)
        ws.cell(row=row_idx, column=4, value=cond_row["sql"]).font = Font(size=10)
        ws.cell(row=row_idx, column=5, value=cond_row["description"]).font = Font(size=10)

        col_ptr = len(_BASE_TITLES) + 1
        for grp, _ in compiled_groups:
            metrics = group_lookup.get(grp, {}).get(cond_row["check_name"], {})
            for metric_key, _ in _METRIC_ORDER:
                val = metrics.get(metric_key, "")
                ws.cell(row=row_idx, column=col_ptr, value=val).font = Font(size=10)
                col_ptr += 1
            # spacer
            col_ptr += 1


def _write_group_table(
    ws,
    *,
    start_row: int,
    group_name: str,
    compiled: List[Tuple[str, pd.DataFrame]],
    conditions: pd.DataFrame,
    start_pop: int | None,
) -> int:
    """Render *compiled* metrics for **one** group starting at *start_row*.

    Returns the next free row index after the table.
    """

    # ------------------------------------------------------------------
    # Header rows – mimic comparison layout but with a single *Current* block
    # ------------------------------------------------------------------

    base_cols = len(_BASE_TITLES)
    block = len(_METRIC_ORDER)
    col_cur = base_cols + 1  # start column for the metric block (1-based)

    # Row 1 (relative to start_row): base descriptive titles
    for cidx, title in enumerate(_BASE_TITLES, start=1):
        cell = ws.cell(row=start_row, column=cidx, value=title)
        cell.font = Font(size=12)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Helper to create a coloured merged-cell header block
    def _mk_block(col_start: int, title: str):
        ws.merge_cells(start_row=start_row,
                       start_column=col_start,
                       end_row=start_row,
                       end_column=col_start + block - 1)
        cell = ws.cell(row=start_row, column=col_start, value=title)
        cell.font = Font(size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _BLOCK_HEADER_FILL

    # Single blue header for CURRENT metrics
    _mk_block(col_cur, "Current")

    # Row 2: metric display names
    row_hdr2 = start_row + 1
    for m_idx, (_, disp) in enumerate(_METRIC_ORDER):
        cell = ws.cell(row=row_hdr2, column=col_cur + m_idx, value=disp)
        cell.font = Font(size=9)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 14

    # ------------------------------------------------------------------
    # Build lookup for metric values (current run only)
    # ------------------------------------------------------------------
    metrics_lookup: Dict[str, Dict[str, Any]] = {}
    for _, df in compiled:
        for _, row in df.iterrows():
            metrics_lookup[row["check_name"]] = row.to_dict()

    # ------------------------------------------------------------------
    # Starting population row
    # ------------------------------------------------------------------
    row_ptr = row_hdr2 + 1
    ws.cell(row=row_ptr, column=5, value="Starting Population").font = Font(size=10, bold=True)
    if start_pop is not None:
        pop_col = len(_BASE_TITLES) + _metric_remaining_offset()
        ws.cell(row=row_ptr, column=pop_col, value=start_pop).font = Font(size=10)

    # ------------------------------------------------------------------
    # Condition rows – one per check
    # ------------------------------------------------------------------
    for _, cond_row in conditions.reset_index().iterrows():
        row_ptr += 1

        ws.cell(row=row_ptr, column=1, value=cond_row["Section"]).font = Font(size=10)
        ws.cell(row=row_ptr, column=2, value=cond_row["Template"]).font = Font(size=10)
        ws.cell(row=row_ptr, column=3, value=cond_row["#"]).font = Font(size=10)
        ws.cell(row=row_ptr, column=4, value=cond_row["sql"]).font = Font(size=10)
        ws.cell(row=row_ptr, column=5, value=cond_row["description"]).font = Font(size=10)

        for m_idx, (mkey, _) in enumerate(_METRIC_ORDER):
            val = metrics_lookup.get(cond_row["check_name"], {}).get(mkey, "")
            ws.cell(row=row_ptr, column=col_cur + m_idx, value=val).font = Font(size=10)

    # Leave a blank line after the table before returning next free row index
    return row_ptr + 2


def _metric_remaining_offset() -> int:
    """Return 1-based offset (within metric block) of 'remaining' column."""
    for idx, (key, _) in enumerate(_METRIC_ORDER, start=1):
        if key == "remaining":
            return idx
    return 1

# ──────────────────────────────────────────────────────────────────────────────
# New comparison writer --------------------------------------------------------
# ──────────────────────────────────────────────────────────────────────────────


def _build_metrics_lookup(compiled: List[Tuple[str, pd.DataFrame]]) -> Dict[str, Dict[str, Any]]:
    """Return mapping check_name -> metric dict for one compiled list."""
    lookup: Dict[str, Dict[str, Any]] = {}
    for _, df in compiled:
        for _, row in df.iterrows():
            lookup[row["check_name"]] = row.to_dict()
    return lookup


def _ordered_union(list_a: Iterable[str], list_b: Iterable[str]) -> List[str]:
    """Return *list_a* followed by any items in *list_b* not already seen, preserving order."""
    seen = set()
    out: List[str] = []
    for item in list_a:
        if item not in seen:
            seen.add(item)
            out.append(item)
    for item in list_b:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def _write_group_comparison_table(
    ws,
    *,
    start_row: int,
    group_name: str,
    compiled_current: List[Tuple[str, pd.DataFrame]],
    compiled_historic: List[Tuple[str, pd.DataFrame]],
    conditions: pd.DataFrame,
    start_pop_current: int | None,
    start_pop_historic: int | None,
    historic_date: str = "",
) -> None:
    """Render side-by-side comparison table for one group."""

    # Column layout constants
    base_cols = len(_BASE_TITLES)
    block = len(_METRIC_ORDER)
    spacer = 1

    # Column indices (1-based)
    col_hist = base_cols + 1
    col_cur = col_hist + block + spacer
    col_diff = col_cur + block + spacer
    col_pct = col_diff + block + spacer

    # Header rows -------------------------------------------------------
    # Row with base titles
    for cidx, title in enumerate(_BASE_TITLES, start=1):
        cell = ws.cell(row=start_row, column=cidx, value=title)
        cell.font = Font(size=12)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Top merged headers for each block
    def _mk_block(col_start, title):
        ws.merge_cells(start_row=start_row, start_column=col_start,
                       end_row=start_row, end_column=col_start + block - 1)
        cell = ws.cell(row=start_row, column=col_start, value=title)
        cell.font = Font(size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _BLOCK_HEADER_FILL

    _mk_block(col_hist, f"Historical {historic_date or ''}")
    _mk_block(col_cur, "Current")
    _mk_block(col_diff, "Δ (Curr-Hist)")
    _mk_block(col_pct, "% Change")

    # Row with metric names
    hdr2 = start_row + 1
    for base_col in [col_hist, col_cur, col_diff, col_pct]:
        for m_idx, (_, disp) in enumerate(_METRIC_ORDER):
            cell = ws.cell(row=hdr2, column=base_col + m_idx, value=disp)
            cell.font = Font(size=9)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = 14


    # Build look-ups and ordered lists
    lookup_cur = _build_metrics_lookup(compiled_current)
    lookup_hist = _build_metrics_lookup(compiled_historic)

    hist_list = list(lookup_hist.keys())
    cur_list = list(lookup_cur.keys())

    # ------------------------------------------------------------------
    # Interleaved row sequence according to spec:
    #   – walk both lists; if names equal → combined row
    #   – else output hist-only row followed by cur-only row
    # ------------------------------------------------------------------
    merged_sequence: list[tuple[str, bool, bool]] = []  # (check_name, has_hist, has_cur)

    i = j = 0
    while i < len(hist_list) or j < len(cur_list):
        h_name = hist_list[i] if i < len(hist_list) else None
        c_name = cur_list[j] if j < len(cur_list) else None

        if h_name is not None and c_name is not None and h_name == c_name:
            merged_sequence.append((h_name, True, True))
            i += 1
            j += 1
        else:
            if h_name is not None:
                merged_sequence.append((h_name, True, False))
                i += 1
            if c_name is not None:
                merged_sequence.append((c_name, False, True))
                j += 1

    # Starting population row ------------------------------------------
    row_ptr = hdr2 + 1
    ws.cell(row=row_ptr, column=5, value="Starting Population").font = Font(size=10, bold=True)

    # Helper to write a value block
    def _write_block(row, col_start, values):
        for m_idx, key in enumerate(_METRIC_ORDER):
            ws.cell(row=row, column=col_start + m_idx, value=values.get(key[0], "")).font = Font(size=10)

    # Starting pop values -> only Remaining metric column is relevant.
    hist_vals = {"remaining": start_pop_historic} if start_pop_historic is not None else {}
    cur_vals = {"remaining": start_pop_current} if start_pop_current is not None else {}

    _write_block(row_ptr, col_hist, hist_vals)
    _write_block(row_ptr, col_cur, cur_vals)

    # diff and pct for start pop
    if start_pop_historic is not None and start_pop_current is not None:
        diff_val = start_pop_current - start_pop_historic
        pct_val = diff_val / start_pop_historic if start_pop_historic else None
        _write_block(row_ptr, col_diff, {"remaining": diff_val})
        _write_block(row_ptr, col_pct, {"remaining": pct_val})

    # Condition rows ----------------------------------------------------
    cond_lookup = conditions.reset_index().set_index('check_name')

    for chk_name, has_hist, has_cur in merged_sequence:
        row_ptr += 1
        # Descriptive columns – fall back to blanks if unknown in config
        if chk_name in cond_lookup.index:
            row_data = cond_lookup.loc[chk_name]
            ws.cell(row=row_ptr, column=1, value=row_data['Section']).font = Font(size=10)
            ws.cell(row=row_ptr, column=2, value=row_data['Template']).font = Font(size=10)
            ws.cell(row=row_ptr, column=3, value=row_data['#']).font = Font(size=10)
            ws.cell(row=row_ptr, column=4, value=row_data['sql']).font = Font(size=10)
            ws.cell(row=row_ptr, column=5, value=row_data['description']).font = Font(size=10)
        else:
            ws.cell(row=row_ptr, column=3, value=chk_name).font = Font(size=10)  # at least display name

        hist_metrics = lookup_hist.get(chk_name, {}) if has_hist else {}
        cur_metrics = lookup_cur.get(chk_name, {}) if has_cur else {}

        _write_block(row_ptr, col_hist, hist_metrics)
        _write_block(row_ptr, col_cur, cur_metrics)

        # Compute diff / pct per metric
        diff_vals = {}
        pct_vals = {}
        for m_key, _ in _METRIC_ORDER:
            h_val = hist_metrics.get(m_key)
            c_val = cur_metrics.get(m_key)
            if h_val is not None and c_val is not None:
                try:
                    diff = c_val - h_val
                    diff_vals[m_key] = diff
                    pct_vals[m_key] = diff / h_val if h_val else ''
                except Exception:
                    pass
        _write_block(row_ptr, col_diff, diff_vals)
        _write_block(row_ptr, col_pct, pct_vals)

    # Borders around full data rectangle
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    last_col = col_pct + block - 1
    for col in range(1, last_col + 1):
        # top
        ws.cell(row=start_row, column=col).border = border
        # bottom
        ws.cell(row=row_ptr, column=col).border = border
    for row in range(start_row, row_ptr + 1):
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=last_col).border = border

    # Freeze panes below the two header rows (i.e., first data row)
    ws.freeze_panes = ws.cell(row=hdr2 + 1, column=1)

    # leave blank row after table
    return
