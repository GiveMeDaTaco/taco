from typing import Dict, List, Any, Optional
from tlptaco.construct_sql.construct_sql import SQLConstructor
from tlptaco.connections.teradata import TeradataHandler
from tlptaco.logging.logging import call_logger, CustomLogger
from tlptaco.eligibility.eligibility import Eligible
from tlptaco.validations.waterfall import WaterfallValidator
from tlptaco.validations.general import BaseValidator
import pandas as pd
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class Waterfall(BaseValidator, WaterfallValidator):
    """
    A class to generate and analyze waterfall reports for campaign eligibility checks.

    Attributes:
        conditions (pd.DataFrame): Conditions for eligibility checks.
        offer_code (str): The offer code.
        campaign_planner (str): The campaign planner.
        lead (str): The lead person.
        waterfall_location (str): The location to save the waterfall report.
        _sqlconstructor (SQLConstructor): An instance of SQLConstructor to build SQL queries.
        _teradata_connection (Optional[TeradataHandler]): The Teradata connection handler.
        _column_names (OrderedDict): Column names for the waterfall report.
        _query_results (Dict[str, List[pd.DataFrame]]): Query results for each identifier.
        _compiled_dataframes (OrderedDict): Compiled dataframes for each identifier.
        _starting_population (Optional[int]): The starting population for the waterfall analysis.
        _combined_df (Optional[pd.DataFrame]): Combined dataframe for the waterfall report.
    """
    _validators = {
        'conditions': WaterfallValidator.validate_conditions,
        'offer_code': WaterfallValidator.validate_offer_code,
        'campaign_planner': WaterfallValidator.validate_campaign_planner,
        'lead': WaterfallValidator.validate_lead,
        'waterfall_location': WaterfallValidator.validate_waterfall_location,
        '_sqlconstructor': WaterfallValidator.validate_sqlconstructor,
        'logger': BaseValidator.validate_logger,
        '_teradata_connection': BaseValidator.validate_teradata_connection
    }

    def __init__(
            self,
            conditions: Dict[str, Dict[str, Any]],
            offer_code: str,
            campaign_planner: str,
            lead: str,
            waterfall_location: str,  # TODO: add to check to see if this directory exists
            sql_constructor: SQLConstructor,
            logger: CustomLogger,
            teradata_connection: Optional[TeradataHandler] = None
    ) -> None:
        """
        Initializes the waterfall class with the provided parameters.

        Args:
            conditions (Dict[str, Dict[str, Any]]): Conditions for eligibility checks.
            offer_code (str): The offer code.
            campaign_planner (str): The campaign planner.
            lead (str): The lead person.
            waterfall_location (str): The location to save the waterfall report.
            sql_constructor (SQLConstructor): An instance of SQLConstructor to build SQL queries.
            teradata_connection (Optional[TeradataHandler]): The Teradata connection handler.
        """
        self.logger = logger
        self.current_date = datetime.now().strftime("%Y-%m-%d %H_%M_%S")
        self._sqlconstructor = sql_constructor
        self._teradata_connection = teradata_connection
        self.offer_code = offer_code
        self.campaign_planner = campaign_planner
        self.lead = lead
        self.waterfall_location = waterfall_location
        self.conditions = conditions

        # prep column names
        self._column_names = OrderedDict({
            'unique_drops': '{identifier} drop if only this drop',
            'increm_drops': '{identifier} drop increm',
            'cumul_drops': '{identifier} drop cumul',
            'regain': '{identifier} regain if no scrub',
            'remaining': '{identifier} remaining'
        })

        # initializing properties
        self._query_results = dict()
        self._compiled_dataframes = OrderedDict()
        self._starting_population = None
        self._combined_df = None

    @classmethod
    def from_eligible(cls, eligibility: Eligible, waterfall_location: str) -> "Waterfall":
        conditions = cls._prepare_conditions(eligibility.conditions)
        offer_code = eligibility.offer_code
        campaign_planner = eligibility.campaign_planner
        lead = eligibility.lead
        sql_constructor = eligibility._sqlconstructor
        logger = eligibility.logger
        teradata_connection = eligibility._teradata_connection

        return cls(conditions, offer_code, campaign_planner, lead, waterfall_location, sql_constructor, logger,
                   teradata_connection)

    @property
    def campaign_planner(self):
        return self._campaign_planner

    @campaign_planner.setter
    def campaign_planner(self, value):
        self._campaign_planner = value

    @property
    def offer_code(self):
        return self._offer_code

    @offer_code.setter
    def offer_code(self, value):
        self._offer_code = value

    @property
    def lead(self):
        return self._lead

    @lead.setter
    def lead(self, value):
        self._lead = value

    @property
    def waterfall_location(self):
        return self._waterfall_location

    @waterfall_location.setter
    def waterfall_location(self, value):
        self._waterfall_location = value

    @classmethod
    def _prepare_conditions(cls, conditions: dict) -> dict:
        """
        Prepares the conditions by transforming them into a dictionary.

        Args:
            conditions (Dict[str, Dict[str, Any]]): Conditions for eligibility checks.

        Returns:
            dict: The prepared conditions as a dictionary.
        """
        result_dict = {}
        for channel, templates in conditions.items():
            for template, checks in templates.items():
                for check in checks:
                    column_name = check.get('column_name', None)
                    description = check.get('description', None)
                    sql = check.get('sql', None)
                    if column_name is not None:
                        modified_description = f'[{template}] {description}' if description else None
                        result_dict[column_name] = {
                            'sql': sql,
                            'description': modified_description
                        }
        return result_dict

    @property
    def conditions(self) -> dict:
        """Getter for conditions."""
        return self._conditions

    @conditions.setter
    def conditions(self, value: Dict[str, Any]) -> None:
        """Setter for conditions."""
        self._conditions = value

    def _save_results(self, identifier: str, data: pd.DataFrame) -> None:
        """
        Saves the results of a query to the _query_results dictionary.

        Args:
            identifier (str): The identifier for the query results.
            data (pd.DataFrame): The data to save.
        """
        # make sure the columns are in order (orderby check number)
        data = data[list(sorted(data.columns, key=lambda x: int(x.split('_')[-1])))]
        if self._query_results.get(identifier) is None:
            self._query_results[identifier] = []
        self._query_results[identifier].append(data)

    @call_logger()
    def _calculate_regain(self) -> None:
        """Calculates the regain SQL and saves the results to the _query_results dictionary."""
        queries = self._sqlconstructor.waterfall.generate_regain_sql()
        # save queries
        self.logger.info(f'{self.__class__}._calculate_regain {queries=}')
        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('regain').format(identifier=identifier)
                df = df.set_index('Index')
                self.logger.info(f"{self.__class__}._calculate_regain {identifier=} {df.to_dict()}")
                self._save_results(identifier, df)

    @call_logger()
    def _calculate_incremental_drops(self) -> None:
        """Calculates the incremental drops SQL and saves the results to the _query_results dictionary."""
        queries = self._sqlconstructor.waterfall.generate_incremental_drops_sql()
        # save queries
        self.logger.info(f'{self.__class__}._calculate_incremental_drops {queries=}')
        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('increm_drops').format(identifier=identifier)
                df = df.set_index('Index')
                self.logger.info(f"{self.__class__}._calculate_incremental_drops {identifier=} {df.to_dict()}")
                self._save_results(identifier, df)

    @call_logger()
    def _calculate_unique_drops(self) -> None:
        """Calculates the unique drops SQL and saves the results to the _query_results dictionary."""
        queries = self._sqlconstructor.waterfall.generate_unique_drops_sql()
        # save queries
        self.logger.info(f'{self.__class__}._calculate_unique_drops {queries=}')
        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('unique_drops').format(identifier=identifier)
                df = df.set_index('Index')
                self.logger.info(f"{self.__class__}._calculate_unique_drops {identifier=} {df.to_dict()}")
                self._save_results(identifier, df)

    @call_logger()
    def _calculate_remaining(self) -> None:
        """Calculates the remaining SQL and saves the results to the _query_results dictionary."""
        queries = self._sqlconstructor.waterfall.generate_remaining_sql()
        # save queries
        self.logger.info(f'{self.__class__}._calculate_remaining {queries=}')
        for identifier, query in queries.items():
            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df['Index'] = self._column_names.get('remaining').format(identifier=identifier)
                df = df.set_index('Index')
                self.logger.info(f"{self.__class__}._calculate_remaining {identifier=} {df.to_dict()}")
                self._save_results(identifier, df)

    def _step1_create_base_tables(self):
        """Creates a summary table for each set of unique identifiers provided, based on the eligibility"""
        queries = self._sqlconstructor.waterfall.generate_unique_identifier_details_sql()
        self.logger.info(f'{self.__class__}._step1_create_base_tables {queries=}')
        for identifier, details in queries.items():
            query = details.get('sql')
            table_name = details.get('table_name')
            collect_query = details.get('collect_stats')
            self.logger.debug(f'--{self.__class__}._step1_create_base_tables\n--SQL for {identifier} stats table\n'
                              f'--{table_name}\n%s;\n--Collect Query\n{collect_query};', query)
            if self._teradata_connection is not None:
                self._teradata_connection.execute_query(query)
                self.logger.info(f"{self.__class__}._step1_create_base_tables created backend table {table_name}")
                self._teradata_connection.tracking.track_table(table_name)
                self.logger.info(f"{self.__class__}._step1_create_base_tables collect statistics on {table_name}")
                self._teradata_connection.execute_query(collect_query)

    def _step2_calculate_stats(self):
        """Grabs the stats from the base tables created in step1 above"""
        queries = self._sqlconstructor.waterfall.generate_all_sql()
        for identifier, query in queries.items():
            self.logger.debug(f'--{self.__class__}._step2_calculate_stats\n--SQL for {identifier} Waterfall Counts\n%s',
                              query)

            unique_drop = self._column_names.get('unique_drops').format(identifier=identifier)
            increm_drop = self._column_names.get('increm_drops').format(identifier=identifier)
            cumul_drop = self._column_names.get('cumul_drops').format(identifier=identifier)
            regain = self._column_names.get('regain').format(identifier=identifier)
            remaining = self._column_names.get('remaining').format(identifier=identifier)

            column_names = {
                self._sqlconstructor.waterfall.generate_regain_sql.__name__: regain,
                self._sqlconstructor.waterfall.generate_incremental_drops_sql.__name__: increm_drop,
                self._sqlconstructor.waterfall.generate_remaining_sql.__name__: remaining,
                self._sqlconstructor.waterfall.generate_unique_drops_sql.__name__: unique_drop
            }

            if self._teradata_connection is not None:
                df = self._teradata_connection.to_pandas(query)
                df = df.T
                df.columns = df.iloc[0]
                df = df[1:]
                # rename the columns
                df = df.rename(columns=column_names, axis=1)
                # make sure the index is in the proper order
                ordered_index = list(sorted(df.index, key=lambda x: int(x.split('_')[-1])))
                df = df.reindex(ordered_index)

                # create cumulative drop column
                self._starting_population = df[increm_drop].values[0] + df[remaining].values[0]
                df[cumul_drop] = self._starting_population - df[remaining]

                df = df[[unique_drop, increm_drop, cumul_drop, regain, remaining]]

                starting_pop_df = pd.DataFrame({
                    unique_drop: 0,
                    increm_drop: 0,
                    cumul_drop: 0,
                    regain: 0,
                    remaining: self._starting_population},
                    index=['main_BA_0']
                )
                df = pd.concat([starting_pop_df, df])
                self.logger.info(f"{self.__class__}._step2_calculate_stats {identifier=}")
                self._compiled_dataframes[identifier] = df

    def _step3_create_excel(self):
        """Creates an excel sheet with the stats gathered in _step2 above"""
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'waterfall'

        # Add header information
        header = f'[[{self.offer_code}] [CP: {self.campaign_planner}] [LEAD: {self.lead}] [DATE: {self.current_date}]'
        ws['A1'] = header
        ws['A1'].font = Font(size=18)

        # Add headers and Starting Population to A2:C2 and C3
        cell = ws.cell(row=2, column=1, value='Checks')
        cell.font = Font(size=12, name='Albany AMT')
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        cell = ws.cell(row=2, column=2, value='Criteria')
        cell.font = Font(size=12, name='Albany AMT')
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        cell = ws.cell(row=2, column=3, value='Description')
        cell.font = Font(size=12, name='Albany AMT')
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        cell = ws.cell(row=3, column=3, value='Starting Population')
        cell.font = Font(size=12, name='Albany AMT')
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # turn conditions into DataFrame and make sure the column order is correct
        conditions_df = pd.DataFrame.from_dict(self.conditions, orient='index')
        conditions_df = conditions_df[['sql', 'description']]

        # put all df's into one list to prep for merge
        all_dfs = [conditions_df] + list(self._compiled_dataframes.values())

        df_combined = None
        for df in all_dfs:
            if df_combined is None:
                df_combined = df
            else:
                df_combined = pd.merge(df_combined, df, left_index=True, right_index=True)
        df_combined.reset_index(drop=False, inplace=True)

        conditions_df = conditions_df.reset_index(drop=False)

        # write the first dataframe values starting from cell A4
        for r_idx, row in enumerate(dataframe_to_rows(conditions_df, index=False, header=False), start=4):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # write other dataframes starting from appropriate columns and add headers
        start_col = 5
        for key, df in self._compiled_dataframes.items():
            for col_num, value in enumerate(df.columns, start=start_col):
                cell = ws.cell(row=2, column=col_num)
                cell.value = value
                cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
                cell.font = Font(size=9, name='Albany AMT')
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                ws.column_dimensions[cell.column_letter].width = 11
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(size=10, name='Albany AMT')
                    cell.number_format = '#,##'
            start_col += len(df.columns) + 1

        channel_name_prev = ''
        header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        col_a_fill = PatternFill(start_color='BBBFAC', end_color='BBBFAC', fill_type='solid')
        blank_fill = PatternFill(start_color='364C40', end_color='364C40', fill_type='solid')

        # loop through the A column, extract the check numbers, and insert a blank row between new channels
        row_modifier = 0  # used to keep count consistent when new rows are inserted
        for row in range(4, len(df_combined) + 5):
            row += row_modifier
            cell = ws.cell(row=row, column=1)
            if cell.value is None:
                continue
            parts = cell.value.split('_')
            channel_name = parts[0]
            check_number = parts[-1]
            if channel_name != channel_name_prev and row != 4:
                ws.insert_rows(row)
                ws.cell(row=row, column=2, value=channel_name.upper())
                # set the background color for this new row
                for col in range(1, start_col):
                    ws.cell(row=row, column=col).fill = header_fill
                row_modifier += 1
            cell.value = check_number
            channel_name_prev = channel_name

        # Apply header fill color to row 2
        for col in range(1, start_col):
            ws.cell(row=2, column=col).fill = header_fill

        # Apply color to column A starting from row 3
        for row in range(3, len(df_combined) + 5):
            ws.cell(row=row, column=1).fill = col_a_fill

        # Apply color to blank rows and columns
        # for row in range(4, len(df_combined) + 5):
        #     if ws.cell(row=row, column=3).value is None:
        #         for col in range(1, start_col):
        #             ws.cell(row=row, column=col).fill = header_fill

        for col in range(1, start_col):
            if ws.cell(row=4, column=col).value is None:
                for row in range(1, len(df_combined) + 5):
                    ws.cell(row=row, column=col).fill = blank_fill

        # set column widths
        ws.column_dimensions['A'].width = 9
        ws.column_dimensions['B'].width = 41
        ws.column_dimensions['C'].width = 51

        # Save the Excel file
        wb.save(f'{self.waterfall_location}/{self.offer_code}_waterfall_{self.current_date}.xlsx')

    def generate_waterfall(self):
        """
        Runs all 3 steps at once
        """
        self._step1_create_base_tables()
        self._step2_calculate_stats()
        self._step3_create_excel()